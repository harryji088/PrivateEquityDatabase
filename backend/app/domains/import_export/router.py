from __future__ import annotations
"""Import/Export REST endpoints with CSV/Excel file upload support."""

import io
import csv
from datetime import datetime, date
from uuid import UUID
from typing import Optional
from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.domains.import_export.models import ImportJob
from app.domains.nav.service import NavService
from app.api.v1.schemas import ApiResponse

router = APIRouter()


@router.post("/nav")
async def import_nav_csv(
    fund_id: UUID = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload a CSV/Excel file for NAV data import."""
    # Validate file type
    if not (file.filename.endswith(".csv") or file.filename.endswith(".xlsx")):
        return ApiResponse(message="Unsupported file type. Use .csv or .xlsx", data=None)

    content = await file.read()

    # Create import job
    job = ImportJob(
        file_name=file.filename,
        file_type=file.filename.rsplit(".", 1)[-1],
        import_type="nav_data",
        status="pending",
    )
    db.add(job)
    await db.flush()

    records = []
    errors = []
    nav_svc = NavService(db)

    if file.filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):  # header is row 1
            try:
                nav_date = date.fromisoformat(row.get("date", "").strip())
                nav_value = float(row.get("nav", "0").strip())
                cum_nav_str = row.get("cumulative_nav", "").strip()
                cumulative_nav = float(cum_nav_str) if cum_nav_str else None

                records.append({
                    "fund_id": fund_id,
                    "date": nav_date,
                    "nav": nav_value,
                    "cumulative_nav": cumulative_nav,
                })
            except (ValueError, KeyError) as e:
                errors.append({"row": i, "message": str(e)})
    elif file.filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                try:
                    row_dict = dict(zip(headers, row))
                    nav_date_val = row_dict.get("date")
                    if isinstance(nav_date_val, datetime):
                        nav_date_val = nav_date_val.date()
                    elif isinstance(nav_date_val, str):
                        nav_date_val = date.fromisoformat(nav_date_val.strip())

                    nav_value = float(row_dict.get("nav", 0))
                    cum_nav_str = row_dict.get("cumulative_nav")
                    cumulative_nav = float(cum_nav_str) if cum_nav_str else None

                    records.append({
                        "fund_id": fund_id,
                        "date": nav_date_val,
                        "nav": nav_value,
                        "cumulative_nav": cumulative_nav,
                    })
                except (ValueError, KeyError) as e:
                    errors.append({"row": row_idx, "message": str(e)})
            wb.close()
        except ImportError:
            return ApiResponse(message="openpyxl not installed", data=None)

    # Process records
    job.total_rows = len(records)
    job.status = "processing"
    await db.flush()

    success = 0
    for rec in records:
        try:
            from app.domains.nav.schemas import NavDataCreate
            await nav_svc.create_nav(NavDataCreate(**rec))
            success += 1
        except Exception as e:
            errors.append({"row": "N/A", "message": str(e)})

    job.success_rows = success
    job.error_rows = len(errors)
    job.error_details = errors
    job.status = "completed" if len(errors) == 0 else "partial"
    job.completed_at = datetime.utcnow()

    return ApiResponse(
        data={
            "job_id": str(job.id),
            "status": job.status,
            "total_rows": job.total_rows,
            "success_rows": success,
            "error_rows": len(errors),
            "errors": errors[:20],  # First 20 errors
        },
        message=f"Import completed: {success}/{len(records)} rows",
    )


@router.get("/jobs/{job_id}/status")
async def get_import_job_status(job_id: UUID, db: AsyncSession = Depends(get_db)):
    """Check import job status."""
    from sqlalchemy import select
    result = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = result.scalar_one_or_none()
    if not job:
        return ApiResponse(message="Job not found", data=None)
    return ApiResponse(data={
        "id": str(job.id),
        "status": job.status,
        "file_name": job.file_name,
        "total_rows": job.total_rows,
        "success_rows": job.success_rows,
        "error_rows": job.error_rows,
        "errors": job.error_details[:10] if job.error_details else [],
    })


@router.get("/template/nav")
async def download_nav_template():
    """Download NAV import CSV template."""
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "nav", "cumulative_nav"])
    writer.writerow(["2024-01-02", "1.0250", "1.0250"])
    writer.writerow(["2024-01-03", "1.0310", "1.0310"])
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=nav_import_template.csv"},
    )
