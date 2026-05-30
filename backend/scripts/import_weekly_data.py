"""
Import weekly quant fund performance data from 点睛业绩放送 Excel files.

Steps:
  1. Parse all 19 weekly Excel files across 7 strategy sheets
  2. Extract and deduplicate fund companies and fund products
  3. Import into database
  4. Build merged time series (weekly returns per fund across all weeks)

Usage:
    python scripts/import_weekly_data.py
"""

import os
import re
import sys
import uuid
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import settings
from app.database import Base
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# ── Configuration ──────────────────────────────────────────────────
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
STRATEGY_SHEET_MAP = {
    "量化选股": "stock_long",
    "市场中性": "market_neutral",
    "500指增": "index_500",
    "1000指增": "index_1000",
    "300指增": "index_300",
    "2000指增": "index_2000",
    "A500指增": "index_a500",
}
STRATEGY_LABELS = {v: k for k, v in STRATEGY_SHEET_MAP.items()}

# Lookup dicts built during import
company_cache = {}   # name -> id
fund_cache = {}      # (company_name, strategy) -> fund_id


def parse_all_files():
    """Parse all weekly Excel files and return structured data."""
    files = sorted([
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.xlsx') and '点睛业绩放送' in f
    ])

    # Data structure: [(week_label, week_end_date, [(fund_record)])]
    all_weeks = []
    # For building fund registry: key=(company_name, strategy) -> {name, strategy, company,...}
    fund_registry = {}
    # For building company registry: name -> {name, size_category}
    company_registry = {}

    print(f"Parsing {len(files)} weekly files...")

    for filepath in files:
        full_path = DATA_DIR / filepath
        wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)

        # Extract week date range from filename
        # e.g., "点睛业绩放送_量化股票策略业绩榜（0105-0109).xlsx"
        match = re.search(r'（(\d{4})-(\d{4})）', filepath)
        if not match:
            match = re.search(r'（(\d{4})-(\d{4})）', filepath.replace('）.xlsx', '）.xlsx'))
        if match:
            week_label = f"{match.group(1)}-{match.group(2)}"

        for sheet_name, strategy_key in STRATEGY_SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Determine which columns exist (量化选股/市场中性 have 14 cols,
            # 指增 sheets have 16-17 cols with extra excess return columns)
            # Read header
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row[1] and '序号' in str(row[1]):
                    header_row = row
                    break

            if not header_row:
                continue

            # Build column mapping
            headers = [str(h) if h else '' for h in header_row]
            col_map = {}
            for i, h in enumerate(headers):
                h_clean = h.strip().replace('\n', '')
                if '序号' in h_clean:
                    col_map['rank'] = i
                elif '策略类型' in h_clean:
                    col_map['strategy'] = i
                elif '基金管理人' in h_clean:
                    col_map['company'] = i
                elif '管理规模' in h_clean:
                    col_map['size'] = i
                elif '净值日期' in h_clean:
                    col_map['date'] = i
                elif '近一周收益率' in h_clean:
                    col_map['weekly_return'] = i
                elif '近一周超额' in h_clean:
                    col_map['weekly_excess'] = i
                elif '今年以来收益率' in h_clean and '年化' not in h_clean:
                    col_map['ytd_return'] = i
                elif '今年以来超额' in h_clean:
                    col_map['ytd_excess'] = i
                elif '今年以来动态回撤' in h_clean and '超额' not in h_clean:
                    col_map['ytd_drawdown'] = i
                elif '成立以来年化收益率' in h_clean:
                    col_map['ann_return'] = i
                elif '成立以来年化波动率' in h_clean:
                    col_map['ann_vol'] = i
                elif '成立以来动态回撤' in h_clean:
                    col_map['max_drawdown'] = i
                elif '成立以来夏普比率' in h_clean or '夏普比率' in h_clean:
                    col_map['sharpe'] = i

            # Parse data rows
            week_records = []
            week_date = None
            for row in ws.iter_rows(min_row=4, values_only=True):
                # Stop at disclaimer or empty row where rank is not numeric
                rank_val = row[col_map.get('rank', 1)] if len(row) > col_map.get('rank', 1) else None
                if rank_val is None or not (isinstance(rank_val, int) or (isinstance(rank_val, float) and rank_val == int(rank_val))):
                    break

                company_name = str(row[col_map['company']]).strip() if row[col_map['company']] else None
                if not company_name:
                    continue

                def safe_float(val):
                    if val is None:
                        return None
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return None

                # Parse date
                date_val = row[col_map['date']]
                if date_val:
                    if isinstance(date_val, datetime):
                        record_date = date_val.date()
                    elif isinstance(date_val, date):
                        record_date = date_val
                    elif isinstance(date_val, str):
                        try:
                            record_date = date.fromisoformat(date_val.strip())
                        except ValueError:
                            continue
                    else:
                        continue
                    if week_date is None:
                        week_date = record_date

                size_category = str(row[col_map['size']]).strip() if row[col_map['size']] else '未知'

                # Register company
                if company_name not in company_registry:
                    company_registry[company_name] = {
                        'name': company_name,
                        'size_category': size_category,
                    }

                # Register fund
                fund_key = (company_name, strategy_key)
                if fund_key not in fund_registry:
                    fund_registry[fund_key] = {
                        'company_name': company_name,
                        'strategy_type': strategy_key,
                        'name': f"{company_name}-{STRATEGY_LABELS[strategy_key]}",
                    }

                record = {
                    'company_name': company_name,
                    'strategy_type': strategy_key,
                    'rank': int(rank_val),
                    'size_category': size_category,
                    'date': record_date,
                    'week_label': week_label,
                    'weekly_return': safe_float(row[col_map.get('weekly_return')]),
                    'weekly_excess': safe_float(row[col_map.get('weekly_excess')]) if 'weekly_excess' in col_map else None,
                    'ytd_return': safe_float(row[col_map.get('ytd_return')]),
                    'ytd_excess': safe_float(row[col_map.get('ytd_excess')]) if 'ytd_excess' in col_map else None,
                    'ytd_drawdown': safe_float(row[col_map.get('ytd_drawdown')]),
                    'ann_return': safe_float(row[col_map.get('ann_return')]),
                    'ann_vol': safe_float(row[col_map.get('ann_vol')]),
                    'max_drawdown': safe_float(row[col_map.get('max_drawdown')]),
                    'sharpe': safe_float(row[col_map.get('sharpe')]),
                }
                week_records.append(record)

            if week_records:
                all_weeks.append({
                    'sheet': sheet_name,
                    'strategy': strategy_key,
                    'week_label': week_label,
                    'week_date': week_date or week_label,
                    'records': week_records,
                })

        wb.close()

    print(f"  Parsed {len(all_weeks)} sheet-weeks")
    print(f"  Found {len(company_registry)} unique companies")
    print(f"  Found {len(fund_registry)} unique fund products (company+strategy)")

    # Count total records
    total_records = sum(len(w['records']) for w in all_weeks)
    print(f"  Total performance records: {total_records}")

    return all_weeks, company_registry, fund_registry


def create_tables(engine):
    """Create weekly_performances table."""
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS weekly_performances (
                id BIGSERIAL PRIMARY KEY,
                fund_id UUID REFERENCES funds(id) ON DELETE CASCADE,
                week_label VARCHAR(20) NOT NULL,
                record_date DATE NOT NULL,
                rank INTEGER,
                weekly_return DOUBLE PRECISION,
                weekly_excess DOUBLE PRECISION,
                ytd_return DOUBLE PRECISION,
                ytd_excess DOUBLE PRECISION,
                ytd_drawdown DOUBLE PRECISION,
                ann_return DOUBLE PRECISION,
                ann_vol DOUBLE PRECISION,
                max_drawdown DOUBLE PRECISION,
                sharpe DOUBLE PRECISION,
                size_category VARCHAR(50),
                created_at TIMESTAMPTZ DEFAULT now(),
                UNIQUE(fund_id, week_label)
            )
        """))
        conn.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_wp_fund_week
            ON weekly_performances(fund_id, week_label)
        """))
        conn.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_wp_strategy_week
            ON weekly_performances(record_date, fund_id)
        """))
        conn.commit()
    print("  weekly_performances table ready")


def import_to_db(all_weeks, company_registry, fund_registry, engine):
    """Import parsed data into database."""
    with Session(engine) as session:
        # 1. Import companies
        print("\nImporting companies...")
        for name, info in company_registry.items():
            from app.domains.companies.models import FundCompany
            existing = session.query(FundCompany).filter_by(name=name).first()
            if existing:
                company_cache[name] = existing.id
            else:
                company = FundCompany(
                    name=name,
                    short_name=name,
                    status='active',
                    description=f"规模: {info['size_category']}",
                )
                session.add(company)
                session.flush()
                company_cache[name] = company.id
        session.commit()
        print(f"  {len(company_cache)} companies in DB")

        # 2. Import funds (one per company+strategy)
        print("Importing fund products...")
        strategy_label_map = {
            'stock_long': '量化选股',
            'market_neutral': '市场中性',
            'index_500': '500指增',
            'index_1000': '1000指增',
            'index_300': '300指增',
            'index_2000': '2000指增',
            'index_a500': 'A500指增',
        }
        from app.domains.funds.models import Fund

        fund_count = 0
        for (company_name, strategy), info in fund_registry.items():
            company_id = company_cache.get(company_name)
            if not company_id:
                continue

            # Check existing fund by company+strategy
            existing = session.query(Fund).filter_by(
                company_id=company_id,
                strategy_type=strategy,
                name=info['name'],
            ).first()

            if existing:
                fund_cache[(company_name, strategy)] = existing.id
            else:
                label = strategy_label_map.get(strategy, strategy)
                fund = Fund(
                    company_id=company_id,
                    name=info['name'],
                    code=None,
                    strategy_type=strategy,
                    inception_date=date(2026, 1, 1),
                    status='active',
                    description=f"{label}策略产品",
                )
                session.add(fund)
                session.flush()
                fund_cache[(company_name, strategy)] = fund.id
                fund_count += 1

        session.commit()
        print(f"  {len(fund_cache)} fund products in DB ({fund_count} new)")

        # 3. Import weekly performance records
        print("Importing weekly performance records...")
        total = 0
        skipped = 0

        for week_data in all_weeks:
            strategy = week_data['strategy']
            week_label = week_data['week_label']

            for rec in week_data['records']:
                fund_id = fund_cache.get((rec['company_name'], strategy))
                if not fund_id:
                    skipped += 1
                    continue

                # Check for duplicate
                from sqlalchemy import text as sql_text
                result = session.execute(
                    sql_text(
                        "SELECT id FROM weekly_performances WHERE fund_id = :fid AND week_label = :wl"
                    ),
                    {"fid": fund_id, "wl": week_label}
                ).fetchone()

                if result:
                    # Update existing
                    session.execute(sql_text("""
                        UPDATE weekly_performances SET
                            rank = :rank, weekly_return = :wr, weekly_excess = :we,
                            ytd_return = :ytd, ytd_excess = :ye, ytd_drawdown = :ydd,
                            ann_return = :ar, ann_vol = :av, max_drawdown = :mdd,
                            sharpe = :sh, size_category = :sc, record_date = :rd
                        WHERE fund_id = :fid AND week_label = :wl
                    """), {
                        "rank": rec['rank'], "wr": rec['weekly_return'],
                        "we": rec['weekly_excess'], "ytd": rec['ytd_return'],
                        "ye": rec['ytd_excess'], "ydd": rec['ytd_drawdown'],
                        "ar": rec['ann_return'], "av": rec['ann_vol'],
                        "mdd": rec['max_drawdown'], "sh": rec['sharpe'],
                        "sc": rec['size_category'], "rd": rec['date'],
                        "fid": fund_id, "wl": week_label,
                    })
                else:
                    session.execute(sql_text("""
                        INSERT INTO weekly_performances
                            (fund_id, week_label, record_date, rank, weekly_return,
                             weekly_excess, ytd_return, ytd_excess, ytd_drawdown,
                             ann_return, ann_vol, max_drawdown, sharpe, size_category)
                        VALUES
                            (:fid, :wl, :rd, :rank, :wr, :we, :ytd, :ye, :ydd,
                             :ar, :av, :mdd, :sh, :sc)
                    """), {
                        "fid": fund_id, "wl": week_label, "rd": rec['date'],
                        "rank": rec['rank'], "wr": rec['weekly_return'],
                        "we": rec['weekly_excess'], "ytd": rec['ytd_return'],
                        "ye": rec['ytd_excess'], "ydd": rec['ytd_drawdown'],
                        "ar": rec['ann_return'], "av": rec['ann_vol'],
                        "mdd": rec['max_drawdown'], "sh": rec['sharpe'],
                        "sc": rec['size_category'],
                    })
                total += 1

        session.commit()
        print(f"  {total} performance records imported ({skipped} skipped)")


def build_time_series(engine):
    """Build and display merged time series per fund across all weeks."""
    print("\n" + "=" * 80)
    print("MERGED TIME SERIES — Weekly Returns by Fund")
    print("=" * 80)

    with engine.connect() as conn:
        # Get all distinct week_labels sorted chronologically
        weeks_result = conn.execute(text("""
            SELECT DISTINCT week_label, MIN(record_date) as rd
            FROM weekly_performances
            GROUP BY week_label
            ORDER BY rd
        """))
        weeks = [(r[0], r[1]) for r in weeks_result]
        week_labels = [w[0] for w in weeks]
        print(f"\nWeeks: {len(weeks)}")
        for wl, wd in weeks:
            print(f"  {wl} ({wd})")

        # Get top funds: those that appear in most weeks
        print("\n--- Top 20 Funds by Coverage (most consistent across weeks) ---")
        top_funds = conn.execute(text("""
            SELECT f.name, fc.name as company, f.strategy_type,
                   COUNT(DISTINCT wp.week_label) as weeks_count
            FROM weekly_performances wp
            JOIN funds f ON f.id = wp.fund_id
            JOIN fund_companies fc ON fc.id = f.company_id
            GROUP BY f.id, f.name, fc.name, f.strategy_type
            ORDER BY weeks_count DESC
            LIMIT 20
        """)).fetchall()

        for row in top_funds:
            print(f"  {row[0]:45s} | {row[1]:15s} | {STRATEGY_LABELS.get(row[2], row[2]):15s} | {row[3]} weeks")

        # Build pivot: fund x week matrix of weekly returns
        print("\n--- Sample Time Series: 量化选股 Top 5 Funds ---")
        funds_result = conn.execute(text("""
            SELECT f.id, f.name, fc.name as company
            FROM weekly_performances wp
            JOIN funds f ON f.id = wp.fund_id
            JOIN fund_companies fc ON fc.id = f.company_id
            WHERE f.strategy_type = 'stock_long'
            GROUP BY f.id, f.name, fc.name
            ORDER BY COUNT(DISTINCT wp.week_label) DESC
            LIMIT 5
        """)).fetchall()

        # Build time series matrix
        for fund_row in funds_result:
            fid, fname, cname = fund_row
            perf_result = conn.execute(text("""
                SELECT week_label, weekly_return, record_date
                FROM weekly_performances
                WHERE fund_id = :fid
                ORDER BY record_date
            """), {"fid": fid}).fetchall()

            returns = {r[0]: r[1] for r in perf_result}
            series = []
            for wl in week_labels:
                ret = returns.get(wl)
                if ret is not None:
                    series.append(f"{ret:+.4f}")
                else:
                    series.append("   ·   ")

            print(f"\n  {fname} ({cname})")
            print(f"  {'  '.join(series)}")

        # Summary stats
        print("\n--- Summary Statistics ---")
        stats = conn.execute(text("""
            SELECT
                f.strategy_type,
                COUNT(DISTINCT f.id) as fund_count,
                COUNT(DISTINCT wp.week_label) as weeks,
                COUNT(*) as total_records,
                ROUND(AVG(wp.weekly_return)::numeric, 6) as avg_weekly_return,
                ROUND(STDDEV(wp.weekly_return)::numeric, 6) as std_weekly_return
            FROM weekly_performances wp
            JOIN funds f ON f.id = wp.fund_id
            GROUP BY f.strategy_type
            ORDER BY f.strategy_type
        """)).fetchall()

        print(f"  {'Strategy':<20s} {'Funds':>6s} {'Weeks':>6s} {'Records':>8s} {'Avg Wk Ret':>12s} {'Std Wk Ret':>12s}")
        print("  " + "-" * 75)
        for row in stats:
            avg_ret = f"{row[4]*100:.2f}%" if row[4] else "N/A"
            std_ret = f"{row[5]*100:.2f}%" if row[5] else "N/A"
            print(f"  {STRATEGY_LABELS.get(row[0], row[0]):20s} {row[1]:>6d} {row[2]:>6d} {row[3]:>8d} {avg_ret:>12s} {std_ret:>12s}")

    # Save time series CSV
    print("\n--- Exporting Merged Time Series to CSV ---")
    csv_path = DATA_DIR.parent / "merged_weekly_returns.csv"
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                f.name as fund_name,
                fc.name as company,
                f.strategy_type,
                wp.week_label,
                wp.record_date,
                wp.weekly_return,
                wp.weekly_excess,
                wp.ytd_return,
                wp.ann_return,
                wp.sharpe,
                wp.size_category
            FROM weekly_performances wp
            JOIN funds f ON f.id = wp.fund_id
            JOIN fund_companies fc ON fc.id = f.company_id
            ORDER BY f.strategy_type, f.name, wp.record_date
        """))

        import csv
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['fund_name', 'company', 'strategy_type', 'week_label',
                             'record_date', 'weekly_return', 'weekly_excess',
                             'ytd_return', 'ann_return', 'sharpe', 'size_category'])
            for row in result:
                writer.writerow(row)

    print(f"  CSV saved to: {csv_path}")


def main():
    engine = create_engine(settings.DATABASE_URL_SYNC, echo=False)

    # Parse all files
    all_weeks, company_registry, fund_registry = parse_all_files()

    # Create table
    print("\nSetting up database...")
    create_tables(engine)

    # Import
    print("\nImporting to database...")
    import_to_db(all_weeks, company_registry, fund_registry, engine)

    # Build time series
    build_time_series(engine)

    print("\n✅ Import complete!")


if __name__ == "__main__":
    main()
