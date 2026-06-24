"""
Import weekly quant fund performance data into SQLite database.
Self-contained — no external database server needed.

Usage:
    python scripts/import_weekly_sqlite.py
"""

import os
import re
import csv
import sys
import sqlite3
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

import openpyxl

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
DB_PATH = DATA_DIR.parent / "cc_data.sqlite3"
CSV_OUTPUT = DATA_DIR.parent / "merged_weekly_returns.csv"

STRATEGY_MAP = {
    "量化选股": "stock_long",
    "市场中性": "market_neutral",
    "500指增": "index_500",
    "1000指增": "index_1000",
    "300指增": "index_300",
    "2000指增": "index_2000",
    "A500指增": "index_a500",
}
STRATEGY_CN = {v: k for k, v in STRATEGY_MAP.items()}


def create_schema(conn):
    """Create database tables."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS fund_companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            size_category TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS funds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL REFERENCES fund_companies(id),
            name TEXT NOT NULL,
            strategy_type TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(company_id, strategy_type)
        );

        CREATE TABLE IF NOT EXISTS weekly_performances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fund_id INTEGER NOT NULL REFERENCES funds(id),
            week_label TEXT NOT NULL,
            record_date TEXT NOT NULL,
            rank INTEGER,
            weekly_return REAL,
            weekly_excess REAL,
            ytd_return REAL,
            ytd_excess REAL,
            ytd_drawdown REAL,
            ytd_excess_drawdown REAL,
            ann_return REAL,
            ann_vol REAL,
            max_drawdown REAL,
            sharpe REAL,
            size_category TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(fund_id, week_label)
        );

        CREATE INDEX IF NOT EXISTS idx_wp_fund ON weekly_performances(fund_id);
        CREATE INDEX IF NOT EXISTS idx_wp_week ON weekly_performances(week_label);
        CREATE INDEX IF NOT EXISTS idx_wp_strategy ON weekly_performances(record_date);
        CREATE INDEX IF NOT EXISTS idx_fund_strategy ON funds(strategy_type);
    """)
    conn.commit()


def parse_all_files():
    """Parse all weekly Excel files."""
    files = sorted([
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.xlsx') and '点睛业绩放送' in f and not f.startswith('~$')
    ])

    all_records = []
    company_set = {}      # name -> size_category
    fund_set = {}         # (company_name, strategy) -> None

    print(f"Parsing {len(files)} weekly files...")

    for fname in files:
        fpath = DATA_DIR / fname
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

        # Extract week label from filename (handles mixed Chinese/ASCII parens)
        match = re.search(r'[（(](\d{4})-(\d{4})[）)]', fname)
        if not match:
            print(f"  WARN: Could not parse date from {fname}")
            wb.close()
            continue
        week_label = f"{match.group(1)}-{match.group(2)}"

        for sheet_name, strategy_key in STRATEGY_MAP.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Find header row
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row[1] and '序号' in str(row[1]):
                    header_row = row
                    break
            if not header_row:
                continue

            headers = [str(h) if h else '' for h in header_row]
            col_map = {}
            for i, h in enumerate(headers):
                h = h.strip()
                if '序号' in h:           col_map['rank'] = i
                elif '策略类型' in h:      col_map['strategy'] = i
                elif '基金管理人' in h:    col_map['company'] = i
                elif '管理规模' in h:      col_map['size'] = i
                elif '净值日期' in h:      col_map['date'] = i
                elif '近一周收益率' in h:   col_map['weekly_return'] = i
                elif '近一周超额' in h:    col_map['weekly_excess'] = i
                elif '今年以来收益率' in h and '年化' not in h:
                    col_map['ytd_return'] = i
                elif '今年以来超额' in h and '动态回撤' not in h:   col_map['ytd_excess'] = i
                elif '今年以来动态回撤' in h and '超额' not in h:
                    col_map['ytd_drawdown'] = i
                elif '成立以来年化收益率' in h:
                    col_map['ann_return'] = i
                elif '成立以来年化波动率' in h:
                    col_map['ann_vol'] = i
                elif '成立以来动态回撤' in h:
                    col_map['max_drawdown'] = i
                elif '成立以来夏普比率' in h or '夏普比率' in h:
                    col_map['sharpe'] = i

            for row in ws.iter_rows(min_row=4, values_only=True):
                rank_val = row[col_map.get('rank', 1)] if len(row) > col_map.get('rank', 1) else None
                if rank_val is None:
                    break
                if not isinstance(rank_val, (int, float)):
                    break

                company_name = str(row[col_map['company']]).strip() if row[col_map['company']] else None
                if not company_name or company_name in ('None', ''):
                    break

                def sf(val):
                    if val is None: return None
                    try: return float(val)
                    except: return None

                date_val = row[col_map.get('date')]
                if isinstance(date_val, datetime):
                    record_date = date_val.date().isoformat()
                elif isinstance(date_val, date):
                    record_date = date_val.isoformat()
                elif isinstance(date_val, str):
                    record_date = date_val.strip()
                else:
                    record_date = None

                if not record_date:
                    continue

                size_cat = str(row[col_map.get('size', 4)]).strip() if col_map.get('size') and row[col_map.get('size')] else '未知'

                company_set[company_name] = size_cat
                fund_set[(company_name, strategy_key)] = None

                # Helper: safely get value from row by column name
                def get_val(key):
                    idx = col_map.get(key)
                    return sf(row[idx]) if idx is not None and idx < len(row) else None

                all_records.append((
                    company_name,
                    strategy_key,
                    week_label,
                    record_date,
                    int(rank_val),
                    get_val('weekly_return'),
                    get_val('weekly_excess'),
                    get_val('ytd_return'),
                    get_val('ytd_excess'),
                    get_val('ytd_drawdown'),
                    None,  # ytd_excess_drawdown — computed later from ytd_excess
                    get_val('ann_return'),
                    get_val('ann_vol'),
                    get_val('max_drawdown'),
                    get_val('sharpe'),
                    size_cat,
                ))

        wb.close()

    print(f"  Parsed {len(all_records)} total performance records")
    print(f"  Found {len(company_set)} unique companies")
    print(f"  Found {len(fund_set)} unique fund products")
    return all_records, company_set, fund_set


def import_to_db(conn, all_records, company_set, fund_set):
    """Import parsed data into SQLite."""
    cur = conn.cursor()

    # Companies
    print("\nImporting companies...")
    company_ids = {}
    for name, size_cat in company_set.items():
        cur.execute(
            "INSERT OR IGNORE INTO fund_companies (name, size_category) VALUES (?, ?)",
            (name, size_cat)
        )
        cur.execute("SELECT id FROM fund_companies WHERE name = ?", (name,))
        company_ids[name] = cur.fetchone()[0]
    conn.commit()
    print(f"  {len(company_ids)} companies")

    # Funds
    print("Importing fund products...")
    fund_ids = {}
    for (company_name, strategy) in fund_set:
        cid = company_ids.get(company_name)
        if not cid:
            continue
        label = STRATEGY_CN.get(strategy, strategy)
        fund_name = f"{company_name}-{label}"
        cur.execute(
            "INSERT OR IGNORE INTO funds (company_id, name, strategy_type) VALUES (?, ?, ?)",
            (cid, fund_name, strategy)
        )
        cur.execute(
            "SELECT id FROM funds WHERE company_id = ? AND strategy_type = ?",
            (cid, strategy)
        )
        fund_ids[(company_name, strategy)] = cur.fetchone()[0]
    conn.commit()
    print(f"  {len(fund_ids)} fund products")

    # Weekly performances
    print("Importing weekly performance records...")
    inserted = 0
    updated = 0
    skipped = 0

    for rec in all_records:
        company_name, strategy_key, week_label, record_date, rank, \
            weekly_return, weekly_excess, ytd_return, ytd_excess, \
            ytd_drawdown, ytd_excess_drawdown, ann_return, ann_vol, max_drawdown, sharpe, size_cat = rec

        fund_id = fund_ids.get((company_name, strategy_key))
        if not fund_id:
            skipped += 1
            continue

        cur.execute(
            "SELECT id FROM weekly_performances WHERE fund_id = ? AND week_label = ?",
            (fund_id, week_label)
        )
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE weekly_performances SET
                    rank=?, weekly_return=?, weekly_excess=?, ytd_return=?,
                    ytd_excess=?, ytd_drawdown=?, ytd_excess_drawdown=?,
                    ann_return=?, ann_vol=?,
                    max_drawdown=?, sharpe=?, size_category=?, record_date=?
                WHERE fund_id=? AND week_label=?
            """, (rank, weekly_return, weekly_excess, ytd_return, ytd_excess,
                  ytd_drawdown, ytd_excess_drawdown, ann_return, ann_vol,
                  max_drawdown, sharpe,
                  size_cat, record_date, fund_id, week_label))
            updated += 1
        else:
            cur.execute("""
                INSERT INTO weekly_performances
                    (fund_id, week_label, record_date, rank, weekly_return,
                     weekly_excess, ytd_return, ytd_excess, ytd_drawdown,
                     ytd_excess_drawdown, ann_return, ann_vol, max_drawdown,
                     sharpe, size_category)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (fund_id, week_label, record_date, rank, weekly_return,
                  weekly_excess, ytd_return, ytd_excess, ytd_drawdown,
                  ytd_excess_drawdown, ann_return, ann_vol, max_drawdown,
                  sharpe, size_cat))
            inserted += 1

    conn.commit()
    print(f"  {inserted} inserted, {updated} updated, {skipped} skipped")

    # Fill excess for market neutral (benchmark = 0, so excess = absolute return)
    cur.execute("""
        UPDATE weekly_performances
        SET weekly_excess = COALESCE(weekly_excess, weekly_return),
            ytd_excess = COALESCE(ytd_excess, ytd_return)
        WHERE fund_id IN (SELECT id FROM funds WHERE strategy_type = 'market_neutral')
    """)
    conn.commit()
    if cur.rowcount > 0:
        print(f"  Filled excess for {cur.rowcount} market neutral records (benchmark=0)")

    # Compute excess for 量化选股 using 中证1000 as benchmark
    compute_stock_long_excess(conn, cur)

    # Compute ytd_excess_drawdown for ALL strategies from ytd_excess
    compute_excess_drawdown(conn, cur)


def compute_stock_long_excess(conn, cur):
    """Compute excess returns for 量化选股 (stock_long) using 中证1000 benchmark.

    weekly_excess = fund_weekly_return - benchmark_weekly_return
    ytd_excess = Π(1 + weekly_excess) - 1  (cumulative compounded)
    """
    import json
    import os

    benchmark_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "benchmark_nav.json"
    )
    if not os.path.exists(benchmark_path):
        # Try relative to CWD
        benchmark_path = "benchmark_nav.json"
    if not os.path.exists(benchmark_path):
        print("  [SKIP] benchmark_nav.json not found, cannot compute stock_long excess")
        return

    with open(benchmark_path, "r") as f:
        bench_data = json.load(f)

    zz1000 = bench_data.get("中证1000")
    if not zz1000:
        print("  [SKIP] 中证1000 not found in benchmark data")
        return

    # Build date -> NAV lookup
    b_map = dict(zip(zz1000["dates"], zz1000["navs"]))

    # Get all weeks sorted by record_date
    cur.execute("""
        SELECT DISTINCT week_label, record_date
        FROM weekly_performances
        ORDER BY record_date
    """)
    weeks = cur.fetchall()

    # Compute benchmark weekly returns
    # For week_i: return = NAV(record_date_i) / NAV(record_date_{i-1}) - 1
    # For the first week: return = NAV(record_date_0) / NAV(first_available_date) - 1
    bench_weekly = {}  # week_label -> benchmark weekly return
    prev_nav = None
    for wl, rd in weeks:
        rd_str = rd.replace("-", "")
        nav = b_map.get(rd_str)
        if nav is None:
            continue
        if prev_nav is None:
            # First week: use the first available benchmark date (start of week)
            first_date = zz1000["dates"][0]
            first_nav = zz1000["navs"][0]
            if first_nav != 0:
                bench_weekly[wl] = nav / first_nav - 1.0
        else:
            bench_weekly[wl] = nav / prev_nav - 1.0
        prev_nav = nav

    # Get stock_long fund IDs
    cur.execute("SELECT id FROM funds WHERE strategy_type = 'stock_long'")
    stock_long_ids = [r[0] for r in cur.fetchall()]

    if not stock_long_ids:
        print("  [SKIP] No stock_long funds found")
        return

    updated = 0
    for fid in stock_long_ids:
        # Get all weekly data for this fund, ordered by date
        cur.execute("""
            SELECT week_label, weekly_return
            FROM weekly_performances
            WHERE fund_id = ?
            ORDER BY record_date
        """, (fid,))
        fund_weeks = {r[0]: r[1] for r in cur.fetchall()}

        # Compute weekly_excess and cumulative ytd_excess
        cum_excess = 0.0
        for wl, rd in weeks:
            if wl not in fund_weeks:
                continue
            weekly_ret = fund_weeks[wl]
            bench_ret = bench_weekly.get(wl)

            if weekly_ret is not None and bench_ret is not None:
                weekly_excess = weekly_ret - bench_ret
                cum_excess = (1.0 + cum_excess) * (1.0 + weekly_excess) - 1.0

                cur.execute("""
                    UPDATE weekly_performances
                    SET weekly_excess = ?, ytd_excess = ?
                    WHERE fund_id = ? AND week_label = ?
                """, (round(weekly_excess, 10), round(cum_excess, 10), fid, wl))
                updated += 1

    conn.commit()
    print(f"  Computed excess for {updated} stock_long records (benchmark=中证1000)")


def compute_excess_drawdown(conn, cur):
    """Compute ytd_excess_drawdown from ytd_excess for ALL funds.

    excess_nav[i]  = 1.0 + ytd_excess[i]
    peak[i]        = max(excess_nav[0..i])
    drawdown[i]    = (excess_nav[i] - peak[i]) / peak[i]
    """
    cur.execute("SELECT id FROM funds")
    fund_ids = [r[0] for r in cur.fetchall()]

    updated = 0
    for fid in fund_ids:
        cur.execute("""
            SELECT week_label, ytd_excess
            FROM weekly_performances
            WHERE fund_id = ? AND ytd_excess IS NOT NULL
            ORDER BY record_date
        """, (fid,))
        rows = cur.fetchall()
        if not rows:
            continue

        peak_excess_nav = -float("inf")
        for wl, ytd_ex in rows:
            excess_nav = 1.0 + ytd_ex
            if excess_nav > peak_excess_nav:
                peak_excess_nav = excess_nav
            if peak_excess_nav != 0:
                dd = (excess_nav - peak_excess_nav) / peak_excess_nav
            else:
                dd = 0.0

            cur.execute("""
                UPDATE weekly_performances
                SET ytd_excess_drawdown = ?
                WHERE fund_id = ? AND week_label = ?
            """, (round(dd, 10), fid, wl))
            updated += 1

    conn.commit()
    print(f"  Computed ytd_excess_drawdown for {updated} records across {len(fund_ids)} funds")


def build_time_series(conn):
    """Build merged time series and export results."""
    cur = conn.cursor()

    # Get all weeks
    cur.execute("""
        SELECT DISTINCT week_label FROM weekly_performances
        ORDER BY week_label
    """)
    week_labels = [r[0] for r in cur.fetchall()]
    print(f"\nTotal weeks: {len(week_labels)}")
    print("Week labels:", ", ".join(week_labels))

    # ── Summary by strategy ──
    print("\n" + "=" * 90)
    print("STRATEGY SUMMARY")
    print("=" * 90)

    cur.execute("""
        SELECT
            f.strategy_type,
            COUNT(DISTINCT f.id) as fund_count,
            COUNT(DISTINCT wp.week_label) as weeks,
            COUNT(*) as total_records,
            ROUND(AVG(wp.weekly_return) * 100, 2) as avg_wk_ret_pct,
            ROUND(AVG(wp.weekly_return) * 100 * 52, 2) as ann_ret_pct,
            ROUND(AVG(wp.ann_return) * 100, 2) as avg_ann_ret_pct,
            ROUND(AVG(wp.sharpe), 2) as avg_sharpe,
            COUNT(DISTINCT wp.week_label || '-' || CAST(wp.fund_id AS TEXT)) as checksum
        FROM weekly_performances wp
        JOIN funds f ON f.id = wp.fund_id
        GROUP BY f.strategy_type
        ORDER BY fund_count DESC
    """)

    print(f"  {'Strategy':<20s} {'Funds':>6s} {'Weeks':>6s} {'Records':>8s} {'Avg WkRet':>10s} {'Ann Est':>10s} {'Avg Sharpe':>10s}")
    print("  " + "-" * 80)
    for row in cur.fetchall():
        strategy, fc, wks, recs, awr, aer, aar, ash, _ = row
        label = STRATEGY_CN.get(strategy, strategy)
        print(f"  {label:<20s} {fc:>6d} {wks:>6d} {recs:>8d} {awr:>8.2f}% {aer:>8.2f}% {ash:>8.2f}")

    # ── Top funds by coverage ──
    print("\n" + "=" * 90)
    print("TOP 25 FUNDS BY COVERAGE (most consistent across weeks)")
    print("=" * 90)

    cur.execute("""
        SELECT f.name, fc.name, f.strategy_type,
               COUNT(DISTINCT wp.week_label) as wks,
               ROUND(AVG(wp.weekly_return) * 100, 2) as avg_ret,
               ROUND(AVG(wp.sharpe), 2) as avg_sh
        FROM weekly_performances wp
        JOIN funds f ON f.id = wp.fund_id
        JOIN fund_companies fc ON fc.id = f.company_id
        GROUP BY f.id
        ORDER BY wks DESC, avg_ret DESC
        LIMIT 25
    """)

    print(f"  {'Fund Name':<40s} {'Company':<15s} {'Strategy':<12s} {'Weeks':>5s} {'Avg Ret%':>9s} {'Avg Sharpe':>10s}")
    print("  " + "-" * 100)
    for row in cur.fetchall():
        fname, cname, strat, wks, avg_ret, avg_sh = row
        label = STRATEGY_CN.get(strat, strat)
        print(f"  {fname:<40s} {cname:<15s} {label:<12s} {wks:>5d} {avg_ret:>7.2f}% {avg_sh:>8.2f}")

    # ── Sample time series: 量化选股 Top 8 ──
    print("\n" + "=" * 90)
    print("TIME SERIES: 量化选股 — Weekly Returns for Top 8 Coverage Funds")
    print("=" * 90)

    cur.execute("""
        SELECT f.id, f.name, fc.name
        FROM weekly_performances wp
        JOIN funds f ON f.id = wp.fund_id
        JOIN fund_companies fc ON fc.id = f.company_id
        WHERE f.strategy_type = 'stock_long'
        GROUP BY f.id
        ORDER BY COUNT(DISTINCT wp.week_label) DESC
        LIMIT 8
    """)
    top_funds = cur.fetchall()

    # Print header
    header = f"  {'Fund':38s}"
    for wl in week_labels:
        header += f" {wl:>10s}"
    print(header)
    print("  " + "-" * (40 + 11 * len(week_labels)))

    for fid, fname, cname in top_funds:
        cur.execute("""
            SELECT week_label, weekly_return FROM weekly_performances
            WHERE fund_id = ? ORDER BY record_date
        """, (fid,))
        returns = {r[0]: r[1] for r in cur.fetchall()}

        series = ""
        for wl in week_labels:
            ret = returns.get(wl)
            if ret is not None:
                series += f" {ret*100:>+8.2f}%"
            else:
                series += f" {'·':>10s}"

        print(f"  {fname:<38s}{series}")

    # ── Time series for 市场中性 ──
    print("\n" + "=" * 90)
    print("TIME SERIES: 市场中性 — Weekly Returns for Top 8 Coverage Funds")
    print("=" * 90)

    cur.execute("""
        SELECT f.id, f.name, fc.name
        FROM weekly_performances wp
        JOIN funds f ON f.id = wp.fund_id
        JOIN fund_companies fc ON fc.id = f.company_id
        WHERE f.strategy_type = 'market_neutral'
        GROUP BY f.id
        ORDER BY COUNT(DISTINCT wp.week_label) DESC
        LIMIT 8
    """)
    top_funds = cur.fetchall()

    print(header)
    print("  " + "-" * (40 + 11 * len(week_labels)))

    for fid, fname, cname in top_funds:
        cur.execute("""
            SELECT week_label, weekly_return FROM weekly_performances
            WHERE fund_id = ? ORDER BY record_date
        """, (fid,))
        returns = {r[0]: r[1] for r in cur.fetchall()}

        series = ""
        for wl in week_labels:
            ret = returns.get(wl)
            if ret is not None:
                series += f" {ret*100:>+8.2f}%"
            else:
                series += f" {'·':>10s}"

        print(f"  {fname:<38s}{series}")

    # ── Export merged CSV ──
    print(f"\nExporting merged time series to CSV: {CSV_OUTPUT}")
    cur.execute("""
        SELECT
            f.name, fc.name, f.strategy_type, wp.week_label,
            wp.record_date, wp.weekly_return, wp.weekly_excess,
            wp.ytd_return, wp.ann_return, wp.sharpe, wp.size_category
        FROM weekly_performances wp
        JOIN funds f ON f.id = wp.fund_id
        JOIN fund_companies fc ON fc.id = f.company_id
        ORDER BY f.strategy_type, f.name, wp.record_date
    """)

    with open(CSV_OUTPUT, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'fund_name', 'company', 'strategy_type', 'week_label',
            'record_date', 'weekly_return', 'weekly_excess',
            'ytd_return', 'ann_return', 'sharpe', 'size_category'
        ])
        writer.writerows(cur.fetchall())

    # Count rows
    with open(CSV_OUTPUT) as f:
        row_count = sum(1 for _ in f) - 1
    print(f"  {row_count} rows exported")


def main():
    full_rebuild = "--full" in sys.argv or "--rebuild" in sys.argv

    print("=" * 60)
    print("  量化私募周度业绩数据导入工具")
    if full_rebuild:
        print("  Mode: FULL rebuild (--full)")
    else:
        print("  Mode: INCREMENTAL (use --full to force rebuild)")
    print("=" * 60)

    if full_rebuild and DB_PATH.exists():
        DB_PATH.unlink()
        print(f"Removed old database: {DB_PATH}")

    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # Create schema (no-op if tables already exist)
    create_schema(conn)

    # Check existing weeks before import (for incremental reporting)
    existing_weeks = set()
    if not full_rebuild:
        cur = conn.cursor()
        try:
            cur.execute("SELECT DISTINCT week_label FROM weekly_performances")
            existing_weeks = {r[0] for r in cur.fetchall()}
        except sqlite3.OperationalError:
            pass  # table doesn't exist yet (first run)

    # Parse all Excel files
    print(f"\nSource directory: {DATA_DIR}")
    all_records, company_set, fund_set = parse_all_files()

    # Report new vs existing weeks
    new_weeks = set()
    for rec in all_records:
        new_weeks.add(rec[2])  # week_label is 3rd element
    fresh_weeks = new_weeks - existing_weeks
    if existing_weeks:
        print(f"  Existing weeks in DB: {len(existing_weeks)}")
        print(f"  New weeks to import:  {len(fresh_weeks)}")
        if fresh_weeks:
            print(f"  New: {', '.join(sorted(fresh_weeks))}")
        else:
            print(f"  ✅ Database is up to date — no new weeks found")
    elif new_weeks:
        print(f"  Weeks to import: {len(new_weeks)}")

    # Import to database
    import_to_db(conn, all_records, company_set, fund_set)

    # Build time series and export
    build_time_series(conn)

    conn.close()

    print(f"\n{'=' * 60}")
    print(f"  Database: {DB_PATH}")
    print(f"  CSV Export: {CSV_OUTPUT}")
    print(f"  ✅ All done!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
